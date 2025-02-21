from flask import Flask, render_template, request, flash, send_file
import requests
from dashboards_with_tables import get_dashboards_with_tables
from openpyxl import Workbook
from io import BytesIO
from datetime import datetime

app = Flask(__name__)
app.secret_key = 'your-secret-key-here'  # Required for flash messages

def create_excel(results):
    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard Queries"
    
    # Add headers
    ws['A1'] = "Dashboard Name"
    ws['B1'] = "Folder"
    ws['C1'] = "Dashboard URL"
    ws['D1'] = "Query"
    
    # Add data
    row = 2
    for dash in results:
        for query in dash['table_queries']:
            ws[f'A{row}'] = dash['name']
            ws[f'B{row}'] = dash['folder']
            ws[f'C{row}'] = dash['url']
            ws[f'D{row}'] = query
            row += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 100
    
    # Create file in memory
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return excel_file

@app.route('/', methods=['GET', 'POST'])
def index():
    results = None
    grafana_url = ''
    api_token = ''
    
    if request.method == 'POST':
        grafana_url = request.form.get('grafana_url', '').rstrip('/')
        api_token = request.form.get('api_token', '')
        
        try:
            results = get_dashboards_with_tables(grafana_url, api_token)
            
            # Check if this is an export request
            if 'export' in request.form:
                excel_file = create_excel(results)
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                return send_file(
                    excel_file,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True,
                    download_name=f'grafana_table_queries_{timestamp}.xlsx'
                )
                
        except Exception as e:
            flash(f"Error: {str(e)}", 'error')
    
    return render_template('index.html', results=results, grafana_url=grafana_url, api_token=api_token)

if __name__ == '__main__':
    # Use this for local development
    app.run(debug=True) 